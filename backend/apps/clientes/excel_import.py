from collections import OrderedDict
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.clientes.models import Cliente
from apps.vehiculos.models import Vehiculo, VehiculoPropietario


# ---------------------------------------------------------------------------
# Definición de columnas del archivo de importación (.xlsx)
# ---------------------------------------------------------------------------
CLIENT_COLUMNS = [
    ('tipo_identificacion', 'Tipo Identificación'),
    ('identificacion', 'Identificación'),
    ('nombre', 'Nombre'),
    ('email', 'Email'),
    ('telefono', 'Teléfono'),
    ('direccion', 'Dirección'),
    ('contifico_id', 'ID Contífico'),
    ('placa', 'Placa'),
    ('vin', 'Chasis/VIN'),
    ('numero_motor', 'Nº Motor'),
    ('marca', 'Marca'),
    ('modelo', 'Modelo'),
    ('anio', 'Año'),
    ('color', 'Color'),
    ('transmision', 'Transmisión'),
    ('combustible', 'Combustible'),
    ('tipo', 'Tipo'),
    ('pais_origen', 'País de Origen'),
    ('kilometraje_actual', 'Kilometraje'),
]

CLIENT_FIELDS = ['tipo_identificacion', 'identificacion', 'nombre', 'email', 'telefono', 'direccion', 'contifico_id']
VEHICLE_FIELDS = ['placa', 'vin', 'numero_motor', 'marca', 'modelo', 'anio', 'color', 'transmision', 'combustible', 'tipo', 'pais_origen', 'kilometraje_actual']

COLUMN_MAP = {key: header for key, header in CLIENT_COLUMNS}

# Alias adicionales aceptados para cada columna (para tolerar variantes de escritura).
HEADER_ALIASES = {
    'identificacion': ['identificacion (cedula/ruc)', 'cedula', 'documento', 'ruc'],
    'nombre': ['nombre o razon social', 'razon social'],
    'numero_motor': ['numero de motor', 'n de motor', 'motor'],
    'vin': ['chasis', 'chasis / vin'],
    'tipo': ['tipo de vehiculo'],
    'combustible': ['tipo de combustible'],
    'pais_origen': ['pais'],
}


def _normalize_header(value):
    import unicodedata
    text = _clean(value).lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if not unicodedata.combining(c))
    text = text.replace('º', '')
    text = ' '.join(text.split())
    return text


def _match_column(header):
    norm = _normalize_header(header)
    for key, canonical in COLUMN_MAP.items():
        if norm == _normalize_header(canonical):
            return key
        for alias in HEADER_ALIASES.get(key, []):
            if norm == _normalize_header(alias):
                return key
    return None


def _clean(value):
    """Normaliza un valor leído de Excel a texto (resuelve tipos muy limpios)."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'true' if value else 'false'
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _coerce_int(value):
    value = _clean(value)
    if not value:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        raise ValueError(f'"{value}" no es un número entero válido.')


def _coerce_choice(value, choices, field_label):
    value = _clean(value).upper()
    if not value:
        return None
    for code, label in choices:
        if value == code:
            return code
        if value in (label.upper(),):
            return code
    raise ValueError(f'"{value}" no es una opción válida para {field_label}.')


def load_rows_from_xlsx(file):
    """Lee el archivo .xlsx y devuelve las filas como dicts (sin encabezados)."""
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    headers_raw = rows[0]
    headers = [_clean(h) for h in headers_raw]

    index_map = {}
    for index, header in enumerate(headers):
        key = _match_column(header)
        if key and key not in index_map:
            index_map[key] = index

    missing = [canonical for key, canonical in COLUMN_MAP.items() if key not in index_map]
    if missing:
        raise ValueError(f'Faltan columnas en el archivo: {", ".join(missing)}')

    data = []
    for row in rows[1:]:
        if all((cell is None or _clean(cell) == '') for cell in row):
            continue
        record = {}
        for key, index in index_map.items():
            raw = row[index] if index < len(row) else None
            record[key] = _clean(raw)
        data.append(record)

    return data, headers


def _validar_cliente(record):
    """Valida y devuelve los datos normalizados de un cliente."""
    nombre = _clean(record.get('nombre'))
    identificacion = _clean(record.get('identificacion'))

    errores = []
    if not nombre:
        errores.append('Falta el nombre o razón social.')
    if not identificacion:
        errores.append('Falta la identificación (cédula/RUC).')

    tipo = 'C'
    if record.get('tipo_identificacion'):
        try:
            tipo = _coerce_choice(record['tipo_identificacion'], Cliente.TIPO_IDENTIFICACION, 'Tipo Identificación')
        except ValueError as e:
            errores.append(str(e))

    if errores:
        raise ValueError(' '.join(errores))

    return {
        'tipo_identificacion': tipo,
        'identificacion': identificacion,
        'nombre': nombre,
        'email': _clean(record.get('email')),
        'telefono': _clean(record.get('telefono')),
        'direccion': _clean(record.get('direccion')),
        'contifico_id': _clean(record.get('contifico_id')),
    }


def _validar_vehiculo(record):
    """Valida y devuelve los datos normalizados de un vehículo o None si no hay datos de vehículo."""
    placa = _clean(record.get('placa'))

    fils_con_en_datos = [
        record.get(k) for k in
        ['placa', 'vin', 'numero_motor', 'marca', 'modelo', 'anio', 'color', 'tipo', 'pais_origen', 'kilometraje_actual']
    ]
    tiene_datos = any(_clean(v) for v in fils_con_en_datos)

    if not tiene_datos:
        return None

    errores = []
    if not placa:
        errores.append('Falta la placa del vehículo.')

    anio = None
    if record.get('anio'):
        try:
            anio = _coerce_int(record['anio'])
        except ValueError as e:
            errores.append(str(e))

    tipo = None
    if record.get('tipo'):
        try:
            tipo = _coerce_choice(record['tipo'], Vehiculo.TipoVehiculo.choices, 'Tipo de Vehículo')
        except ValueError as e:
            errores.append(str(e))
    else:
        tipo = Vehiculo.TipoVehiculo.AUTOMOVIL

    transmision = None
    if record.get('transmision'):
        try:
            transmision = _coerce_choice(record['transmision'], Vehiculo.TipoTransmision.choices, 'Transmisión')
        except ValueError as e:
            errores.append(str(e))
    else:
        transmision = Vehiculo.TipoTransmision.MANUAL

    combustible = None
    if record.get('combustible'):
        try:
            combustible = _coerce_choice(record['combustible'], Vehiculo.TipoCombustible.choices, 'Combustible')
        except ValueError as e:
            errores.append(str(e))
    else:
        combustible = Vehiculo.TipoCombustible.GASOLINA

    kilometraje = 0
    if record.get('kilometraje_actual'):
        try:
            kilometraje = _coerce_int(record['kilometraje_actual'])
        except ValueError as e:
            errores.append(str(e))

    pais_origen = _clean(record.get('pais_origen')) or 'EC'
    from django_countries.data import COUNTRIES
    if pais_origen.upper() not in COUNTRIES:
        errores.append(f'"{pais_origen}" no es un código de país válido (ISO 3166).')

    if errores:
        raise ValueError(' '.join(errores))

    return {
        'placa': placa.upper(),
        'vin': _clean(record.get('vin')),
        'numero_motor': _clean(record.get('numero_motor')),
        'marca': _clean(record.get('marca')),
        'modelo': _clean(record.get('modelo')),
        'anio': anio,
        'color': _clean(record.get('color')),
        'transmision': transmision,
        'combustible': combustible,
        'tipo': tipo,
        'pais_origen': pais_origen.upper(),
        'kilometraje_actual': kilometraje,
    }


def _agrupar_por_cliente(rows):
    """Agrupa filas por identificación del cliente, manteniendo el número de fila original."""
    grupos = OrderedDict()
    for numero_fila, record in rows:
        clave = record.get('identificacion', '').strip()
        if clave not in grupos:
            grupos[clave] = []
        grupos[clave].append((numero_fila, record))
    return grupos


def importar_clientes_desde_xlsx(file, empresa_id):
    """
    Procesa un archivo .xlsx de clientes + vehículos de forma tolerante a fallos.

    - Agrupa las filas por identificación del cliente (mismo cliente con varios vehículos).
    - Cada fila o grupo fallido se registra sin detener el procesamiento.
    - Devuelve un resumen con totales y errores por fila.

    Devuelve un dict con: total, exitosos, errores.
    """
    resultado = {
        'total': 0,
        'exitosos': 0,
        'errores': [],
    }

    rows, _headers = load_rows_from_xlsx(file)
    resultado['total'] = len(rows)

    # Registramos el número de fila real (sumando 1 por la fila de encabezados).
    filas = [(idx + 2, record) for idx, record in enumerate(rows)]
    grupos = _agrupar_por_cliente(filas)

    for identificacion, filas_grupo in grupos.items():
        numeros_fila = [fila[0] for fila in filas_grupo]
        primer_numero = numeros_fila[0]

        try:
            datos_cliente = _validar_cliente(filas_grupo[0][1])
        except Exception as e:
            resultado['errores'].append({
                'fila': primer_numero,
                'identificacion': identificacion,
                'datos': filas_grupo[0][1],
                'motivo': str(e),
            })
            continue

        try:
            cliente, _creado = Cliente.objects.update_or_create(
                empresa_id=empresa_id,
                identificacion=datos_cliente['identificacion'],
                defaults={
                    'tipo_identificacion': datos_cliente['tipo_identificacion'],
                    'nombre': datos_cliente['nombre'],
                    'email': datos_cliente['email'],
                    'telefono': datos_cliente['telefono'],
                    'direccion': datos_cliente['direccion'],
                    'contifico_id': datos_cliente['contifico_id'],
                    'is_active': True,
                },
            )
        except Exception as e:
            resultado['errores'].append({
                'fila': primer_numero,
                'identificacion': identificacion,
                'datos': filas_grupo[0][1],
                'motivo': f'No se pudo crear/actualizar el cliente: {e}',
            })
            continue

        grupo_ok = True
        hubo_filas_validas = False
        for numero_fila, record in filas_grupo:
            try:
                datos_vehiculo = _validar_vehiculo(record)
                if not datos_vehiculo:
                    hubo_filas_validas = True
                    continue

                vehiculo, _vcreado = Vehiculo.objects.update_or_create(
                    placa=datos_vehiculo['placa'],
                    defaults={
                        'vin': datos_vehiculo['vin'],
                        'numero_motor': datos_vehiculo['numero_motor'],
                        'marca': datos_vehiculo['marca'],
                        'modelo': datos_vehiculo['modelo'],
                        'anio': datos_vehiculo['anio'],
                        'color': datos_vehiculo['color'],
                        'transmision': datos_vehiculo['transmision'],
                        'combustible': datos_vehiculo['combustible'],
                        'tipo': datos_vehiculo['tipo'],
                        'pais_origen': datos_vehiculo['pais_origen'],
                        'kilometraje_actual': datos_vehiculo['kilometraje_actual'],
                        'is_active': True,
                    },
                )
                vehiculo.empresas.add(empresa_id)

                VehiculoPropietario.objects.update_or_create(
                    vehiculo=vehiculo,
                    cliente=cliente,
                    defaults={'es_actual': True, 'fecha_fin': None},
                )
                hubo_filas_validas = True
            except Exception as e:
                grupo_ok = False
                resultado['errores'].append({
                    'fila': numero_fila,
                    'identificacion': identificacion,
                    'datos': record,
                    'motivo': str(e),
                })

        if grupo_ok and hubo_filas_validas:
            resultado['exitosos'] += 1

    return resultado


# ---------------------------------------------------------------------------
# Reporte de errores (.xlsx)
# ---------------------------------------------------------------------------
def generar_reporte_errores_xlsx(errores):
    """Genera un archivo .xlsx con el detalle exacto de los registros que fallaron."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Errores de Importación'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='b91c1c', end_color='b91c1c', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='e5e7eb'),
        right=Side(style='thin', color='e5e7eb'),
        top=Side(style='thin', color='e5e7eb'),
        bottom=Side(style='thin', color='e5e7eb'),
    )

    fecha_hora = timezone.localtime().strftime('%d/%m/%Y %H:%M')
    titulo = ws.cell(row=1, column=1, value=f'Reporte de errores de importación - Generado: {fecha_hora}')
    titulo.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMN_MAP) + 2)

    headers = ['Fila', 'Motivo del Fallo'] + [header for _, header in CLIENT_COLUMNS]
    header_row = 3

    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    current_row = header_row + 1
    for error in errores:
        motivo = error.get('motivo') or ''
        ws.cell(row=current_row, column=1, value=error.get('fila') or '')
        ws.cell(row=current_row, column=2, value=motivo)
        datos = error.get('datos') or {}
        for col_idx, (key, _header) in enumerate(CLIENT_COLUMNS, start=3):
            ws.cell(row=current_row, column=col_idx, value=datos.get(key) or '')
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col_idx).alignment = cell_alignment
            ws.cell(row=current_row, column=col_idx).border = thin_border
        current_row += 1

    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Plantilla oficial de importación (.xlsx)
# ---------------------------------------------------------------------------
def generar_plantilla_xlsx():
    """Genera la plantilla oficial de importación con los encabezados exactos en la primera fila y filas de ejemplo."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Importación Clientes'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='e5e7eb'),
        right=Side(style='thin', color='e5e7eb'),
        top=Side(style='thin', color='e5e7eb'),
        bottom=Side(style='thin', color='e5e7eb'),
    )

    # Fila 1: encabezados exactos esperados por el backend.
    header_row = 1
    for col_idx, (key, header_text) in enumerate(CLIENT_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ejemplo = [
        'C', '1712345678', 'Juan Pérez', 'juan@correo.com', '0991234567', 'Av. Siempre Viva 123',
        '', 'PBA1234', 'VIN123456789', 'MOTOR123', 'Toyota', 'Corolla', '2020', 'Rojo',
        'M', 'GAS', 'AUTO', 'EC', '45000',
    ]
    ejemplo_veh2 = [
        'C', '1712345678', 'Juan Pérez', '', '', '',
        '', 'PBA5678', 'VIN987654321', '', 'Toyota', 'Hilux', '2019', 'Plata',
        'M', 'DIE', 'CAMN', 'EC', '82000',
    ]

    example_fill = PatternFill(start_color='fef9c3', end_color='fef9c3', fill_type='solid')
    for fila, datos in ((2, ejemplo), (3, ejemplo_veh2)):
        for col_idx, value in enumerate(datos, start=1):
            cell = ws.cell(row=fila, column=col_idx, value=value or '')
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.fill = example_fill

    # Nota orientativa fuera del área de datos (columna siguiente a los encabezados).
    nota_col = len(CLIENT_COLUMNS) + 2
    ws.cell(row=1, column=nota_col, value='Nota:')
    nota_cell = ws.cell(
        row=1, column=nota_col + 1,
        value=('Si un cliente tiene varios vehículos, repite sus datos personales en cada fila '
               'agrupadas por su número de identificación. Borra las filas de ejemplo antes de importar.'),
    )
    nota_cell.font = Font(size=10, italic=True, color='6b7280')

    for col_idx, (key, _header) in enumerate(CLIENT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
