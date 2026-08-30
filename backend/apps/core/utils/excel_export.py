from io import BytesIO
from typing import List, Tuple, Callable, Optional

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelExportConfig:
    def __init__(
        self,
        title: str,
        filename: str,
        headers: List[Tuple[str, float]],
        row_builder: Optional[Callable] = None,
        empresa=None,
        taller=None,
        usuario: Optional[str] = None,
    ):
        self.title = title
        self.filename = filename
        self.headers = headers
        self.row_builder = row_builder
        self.empresa = empresa
        self.taller = taller
        self.usuario = usuario


class ExcelExportService:
    def __init__(self, config: ExcelExportConfig, queryset):
        self.config = config
        self.queryset = queryset

    def _build_header_rows(self) -> List[dict]:
        rows = []
        empresa = getattr(self.config, 'empresa', None)
        taller = getattr(self.config, 'taller', None)

        if empresa:
            nombre_empresa = getattr(empresa, 'razon_social', '') or getattr(empresa, 'nombre_comercial', '') or ''
            if nombre_empresa:
                rows.append({'text': nombre_empresa, 'bold': True, 'size': 14})
            ruc = getattr(empresa, 'ruc', '') or ''
            if ruc:
                rows.append({'text': f'RUC: {ruc}', 'size': 11})
            telefono = getattr(empresa, 'telefono', '') or ''
            if telefono:
                rows.append({'text': f'Tel: {telefono}', 'size': 11})
            if taller:
                direccion = getattr(taller, 'direccion', '') or ''
                if direccion:
                    rows.append({'text': f'Dirección: {direccion}', 'size': 11})

        from django.utils import timezone
        fecha_hora = timezone.localtime().strftime('%d/%m/%Y %H:%M')
        rows.append({'text': f'{self.config.title} - Generado: {fecha_hora}', 'bold': True, 'size': 12})

        usuario = getattr(self.config, 'usuario', '') or ''
        if usuario:
            rows.append({'text': f'Usuario: {usuario}', 'size': 11})

        return rows

    def _build_data_rows(self):
        rows = []
        if self.config.row_builder:
            for instance in self.queryset:
                rows.append(self.config.row_builder(instance))
        else:
            for instance in self.queryset:
                rows.append([str(instance)])
        return rows

    def generate_response(self) -> HttpResponse:
        wb = Workbook()
        ws = wb.active
        ws.title = self.config.title

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        info_font = Font(bold=True, size=14)
        info_alignment = Alignment(horizontal='left', vertical='center')

        cell_alignment = Alignment(vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='e5e7eb'),
            right=Side(style='thin', color='e5e7eb'),
            top=Side(style='thin', color='e5e7eb'),
            bottom=Side(style='thin', color='e5e7eb'),
        )

        current_row = 1
        for info in self._build_header_rows():
            cell = ws.cell(row=current_row, column=1, value=info['text'])
            cell.font = info_font if info.get('bold') else Font(size=info.get('size', 11))
            cell.alignment = info_alignment
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(self.config.headers))
            current_row += 1

        if self._build_header_rows():
            current_row += 1

        for col_idx, (header_text, _) in enumerate(self.config.headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        current_row += 1

        for row_data in self._build_data_rows():
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value or '')
                cell.alignment = cell_alignment
                cell.border = thin_border
            current_row += 1

        for col_idx, (_, width_inches) in enumerate(self.config.headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(width_inches * 8, 12)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.config.filename}"'
        return response
